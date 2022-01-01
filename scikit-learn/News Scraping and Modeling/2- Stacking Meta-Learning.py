### Libs

import re, os
import string
import pandas as pd
import numpy as np
from numpy import std
from numpy import mean
from functools import lru_cache

import spacy
import sklearn

from sklearn.pipeline import Pipeline
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import HashingVectorizer
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.feature_extraction.text import TfidfVectorizer

from sklearn.model_selection import GridSearchCV
from sklearn.model_selection import RandomizedSearchCV
from sklearn.model_selection import train_test_split
from sklearn.model_selection import cross_val_score
from sklearn.model_selection import cross_validate # to use sevrela validation metrics
from sklearn import metrics
from sklearn.utils import shuffle
from sklearn.model_selection import ShuffleSplit
from sklearn.model_selection import RepeatedStratifiedKFold

from sklearn import svm
from sklearn.svm import SVC
from sklearn.naive_bayes import GaussianNB
from sklearn.linear_model import LogisticRegression
from sklearn.naive_bayes import MultinomialNB
from sklearn.linear_model import SGDClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.tree import DecisionTreeClassifier
from sklearn.ensemble import StackingClassifier
from sklearn.neighbors import KNeighborsClassifier

from matplotlib import pyplot

from joblib import dump, load

### Data
news = pd.read_excel()

***

news = news.sample(frac=1, axis=1).sample(frac=1).reset_index(drop=True)

print(news['usefull'].value_counts())
print('Data shape: ', news.shape)
news.head()

### NLP Processing

punctuations = string.punctuation
nlp = spacy.load("de_core_news_sm")
stop_words = spacy.lang.de.stop_words.STOP_WORDS

@lru_cache(maxsize=10000)
def spacy_tokenizer(sentence):
    
    mytokens = []
    
#     # configuration
#     mytokens = []
#     punctuations = string.punctuation
#     nlp = spacy.load("de_core_news_sm")
#     stop_words = spacy.lang.de.stop_words.STOP_WORDS
    
#     # Remove trailling and overflow white spaces
#     sentence = re.sub("\s\s+" , " ", sentence.strip()) #takes too much time
    
    # Lemmatizing each token and converting each token into lowercase
    mytokens = [mytokens.append(word.lemma_) or word.lemma_ for word in nlp(re.sub("\s\s+" , " ", sentence.strip()))] # it automatically lowercase the letters

    # Removing stop words and punctuation
    mytokens = [ word for word in mytokens if word not in stop_words and word not in punctuations ]

    return mytokens

### Test
#%%time
#sentences = sentences.tolist()
#df_tmp = list(map(lambda sentence : spacy_tokenizer(sentence), sentences))
#df_tmp[:3]

### Stacking

# get a stacking ensemble of models
def get_stacking():
    # define the base models
    level0 = list()
    #level0.append(('lr'  , LogisticRegression()))
    #level0.append(('knn' , KNeighborsClassifier()))
    #level0.append(('cart', DecisionTreeClassifier()))
    level0.append(('mlp', MLPClassifier())) # MLPClassifier(alpha=0.01, hidden_layer_sizes=11, max_iter=500, random_state=2)
    level0.append(('svm', SVC()))           # SVM works relatively well when there is a clear margin of separation between classes. SVM is more effective in high dimensional spaces. SVM is effective in cases where the number of dimensions is greater than the number of samples. SVM is relatively memory efficient.
    #level0.append(('bayes', GaussianNB()))
    
    # define meta learner model
    level1 = LogisticRegression()          # LogisticRegression(C=0.5, class_weight='balanced', max_iter=3000, solver='saga')
    # define the stacking ensemble
    model = StackingClassifier(estimators=level0, final_estimator=level1, cv=3)
    return model

# get a list of models to evaluate
def get_models():
    models             = dict()
    
    # LogisticRegression(C=0.5, class_weight='balanced', max_iter=3000, solver='saga')
    models['lr']       = LogisticRegression() 
    
    #models['knn']     = KNeighborsClassifier()
    #models['cart']    = DecisionTreeClassifier()
    
    # MLPClassifier(alpha=0.01, hidden_layer_sizes=11, max_iter=500, random_state=2)
    models['mlp']      = MLPClassifier() # can be applied to complex non-linear problems. Works well with large input data. Provides quick predictions after training. The same accuracy ratio can be achieved even with smaller data.
    
    # SVC(C=0.1, gamma=1, kernel='linear')
    models['svm']      = SVC() # SVM works relatively well when there is a clear margin of separation between classes. SVM is more effective in high dimensional spaces. SVM is effective in cases where the number of dimensions is greater than the number of samples. SVM is relatively memory efficient.
    
    #models['bayes']   = GaussianNB()
    models['stacking'] = get_stacking()
    return models
 
# evaluate a give model using cross-validation
def evaluate_model(model, X, y):
    cv      = RepeatedStratifiedKFold(n_splits=5, n_repeats=3, random_state=71)
    scoring = ['accuracy', 'precision_macro', 'recall_macro', 'f1_macro']
    scores  = cross_val_score(model, X, y, scoring=scoring, cv=cv, n_jobs=-1, error_score='raise')
    return scores
  
  
### Processing

%%time

X         = news['text'].astype(str) 
ylabels   = news['usefull'].astype(int) 
X_train, X_test, y_train, y_test = train_test_split(X, ylabels, test_size=0.25, shuffle=True, stratify=ylabels)

count_vect = CountVectorizer(tokenizer=spacy_tokenizer, encoding='utf-8', decode_error='ignore', strip_accents='unicode', lowercase=True, analyzer='word', ngram_range=(2, 2), max_features=55000)# tokenizer = spacy_tokenizer  
count_vect = count_vect.fit(X_train)
X_train    = count_vect.transform(X_train)

tfidf_transformer = TfidfTransformer(use_idf=True)
tfidf_transformer = tfidf_transformer.fit(X_train)
X_train           = tfidf_transformer.transform(X_train)

X_test = count_vect.transform(X_test)
X_test = tfidf_transformer.transform(X_test)

%%time

# Save Count Vectorizer after fiting with X_train
dump(count_vect, r'\countvect_v0.joblib') 

# Load the Count Vectorizer
# cv_test = load(r'\countvect_v0.joblib')
# test = cv_test.transform(['mk is the best', 'hello mk'])
# test = tfidf_transformer.transform(test)

dump(tfidf_transformer, r'\tfidftransformer_v0.joblib') 

### Evaluating Models

%%time

# get the models to evaluate
models = get_models()
# evaluate the models and store results
results, names = list(), list()
for name, model in models.items():
	scores = evaluate_model(model, X_train, y_train)
	results.append(scores)
	names.append(name)
	print('>%s %.3f (%.3f)' % (name, mean(scores), std(scores)))
# plot model performance for comparison
pyplot.boxplot(results, labels=names, showmeans=True) # green==mean orange=median # the dimensions of the box is determined by the Standard deviation
# In statistics, the standard deviation is a measure of the amount of variation or dispersion of a set of values. 
# A low standard deviation indicates that the values tend to be close to the mean of the set, while a high standard deviation indicates that the values are spread out over a wider range.
# The mean (average) of a data set is found by adding all numbers in the data set and then dividing by the number of values in the set. 
# The median is the middle value when a data set is ordered from least to greatest.
# The circles are ouliers
# for more info https://stackoverflow.com/questions/17725927/boxplots-in-matplotlib-markers-and-outliers
pyplot.show()

### Fit & Save, Predict & Export

%%time

text_clf  = get_stacking()
model     = text_clf.fit(X_train, y_train)

# Save model
dump(model, r'\model_v0.joblib') 

# Load model
#clf = load(r'\model_v0.joblib')

%%time

# Since we shuffle the data, this is the only way to get the text and it's predicted usefullness in order
test = count_vect.transform(X)
test = tfidf_transformer.transform(test)
predicted = model.predict(test)

news['Prediction'] = list(predicted)
news['usefull']   = news['usefull'].astype(int)
cols = ['url', 'title', 'text', 'usefull', 'Prediction']
news=news[cols]
news.head()

news.to_excel(r'\Results.xlsx',  index=False)

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

wb = load_workbook(r'\Results.xlsx')
ws = wb.active

redFill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')

for row in range(len(news)):
    if ws['D'+str(row+2)].value != ws['E'+str(row+2)].value:
        ws['E'+str(row+2)].fill = redFill
        print(row)

wb.save(r'\Results.xlsx')
